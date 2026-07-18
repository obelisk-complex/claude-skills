#!/usr/bin/env python3
"""
Generate a BOM (Bill of Materials) spreadsheet from a JSON parts list.

Usage:
    python generate_bom.py input.json output.xlsx

Input JSON format:
[
  {
    "ref_des": "R1 R2 R3",
    "value": "10kΩ",
    "description": "Resistor, 10kΩ, 1%, 0402",
    "footprint": "Resistor_SMD:R_0402_1005Metric",
    "package": "0402",
    "mpn": "RC0402FR-0710KL",
    "manufacturer": "Yageo",
    "qty": 3,
    "supplier": "DigiKey",
    "supplier_pn": "311-10.0KLRCT-ND",
    "unit_price_1": 0.01,
    "unit_price_100": 0.005,
    "alt1_mpn": "CRCW040210K0FKED",
    "alt1_mfr": "Vishay",
    "alt1_compat": "Drop-in",
    "alt1_notes": "Same spec, different manufacturer",
    "alt2_mpn": "ERJ-2RKF1002X",
    "alt2_mfr": "Panasonic",
    "alt2_compat": "Drop-in",
    "alt2_notes": "Same spec, different manufacturer"
  }
]

Fields alt1_* and alt2_* are optional. All other fields are required.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# Column definitions: (header, json_key, width)
COLUMNS = [
    ("Ref Des", "ref_des", 14),
    ("Value", "value", 12),
    ("Description", "description", 30),
    ("Footprint", "footprint", 28),
    ("Package", "package", 10),
    ("MPN (Primary)", "mpn", 22),
    ("Manufacturer", "manufacturer", 16),
    ("Qty", "qty", 6),
    ("Supplier", "supplier", 12),
    ("Supplier PN", "supplier_pn", 20),
    ("Unit Price (1pc)", "unit_price_1", 14),
    ("Unit Price (100pc)", "unit_price_100", 15),
    ("Alt 1 MPN", "alt1_mpn", 22),
    ("Alt 1 Mfr", "alt1_mfr", 14),
    ("Alt 1 Compat", "alt1_compat", 14),
    ("Alt 1 Notes", "alt1_notes", 25),
    ("Alt 2 MPN", "alt2_mpn", 22),
    ("Alt 2 Mfr", "alt2_mfr", 14),
    ("Alt 2 Compat", "alt2_compat", 14),
    ("Alt 2 Notes", "alt2_notes", 25),
]

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=10)
PRICE_FORMAT = '"$"#,##0.000'
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ZEBRA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def generate_bom(parts: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    # Freeze top row and first column
    ws.freeze_panes = "B2"

    # Write headers
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        # Alt columns get green header
        if header.startswith("Alt"):
            cell.fill = ALT_HEADER_FILL
        else:
            cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    for row_idx, part in enumerate(parts, start=2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = part.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER

            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL

            # Price formatting
            if key in ("unit_price_1", "unit_price_100") and isinstance(value, (int, float)):
                cell.number_format = PRICE_FORMAT

            # Centre-align short columns
            if key in ("qty", "package", "alt1_compat", "alt2_compat"):
                cell.alignment = Alignment(horizontal="center")

    # Add summary row
    summary_row = len(parts) + 3
    ws.cell(row=summary_row, column=1, value="Total unique line items:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(parts))

    total_qty = sum(p.get("qty", 0) for p in parts)
    ws.cell(row=summary_row + 1, column=1, value="Total component count:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=total_qty)

    # Auto-filter on headers
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(parts) + 1}"

    # Add JLCPCB assembly sheet
    ws_jlc = wb.create_sheet("JLCPCB BOM")
    jlc_headers = ["Comment", "Designator", "Footprint", "LCSC Part #"]
    for col_idx, header in enumerate(jlc_headers, start=1):
        cell = ws_jlc.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, part in enumerate(parts, start=2):
        ws_jlc.cell(row=row_idx, column=1, value=part.get("value", ""))
        ws_jlc.cell(row=row_idx, column=2, value=part.get("ref_des", ""))
        # Extract just the package from the footprint
        ws_jlc.cell(row=row_idx, column=3, value=part.get("package", ""))
        ws_jlc.cell(row=row_idx, column=4, value=part.get("lcsc_pn", ""))

    ws_jlc.column_dimensions["A"].width = 20
    ws_jlc.column_dimensions["B"].width = 20
    ws_jlc.column_dimensions["C"].width = 15
    ws_jlc.column_dimensions["D"].width = 15

    # Add CPL template sheet
    ws_cpl = wb.create_sheet("JLCPCB CPL (template)")
    cpl_headers = ["Designator", "Val", "Package", "Mid X", "Mid Y", "Rotation", "Layer"]
    for col_idx, header in enumerate(cpl_headers, start=1):
        cell = ws_cpl.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    note_row = 2
    ws_cpl.cell(
        row=note_row, column=1,
        value="Export position file from KiCad (File → Fabrication Outputs → Component Placement) "
              "and paste data here, or upload the .pos file directly to JLCPCB."
    ).font = Font(italic=True, color="666666")
    ws_cpl.merge_cells(f"A{note_row}:G{note_row}")

    wb.save(output_path)
    print(f"BOM written to {output_path}")
    print(f"  {len(parts)} line items, {total_qty} total components")
    print(f"  Sheets: BOM, JLCPCB BOM, JLCPCB CPL (template)")


def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} input.json output.xlsx")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = sys.argv[2]

    if not input_path.exists():
        print(f"Error: {input_path} not found")
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        parts = json.load(f)

    if not isinstance(parts, list):
        print("Error: JSON must be an array of part objects")
        sys.exit(1)

    generate_bom(parts, output_path)


if __name__ == "__main__":
    main()
